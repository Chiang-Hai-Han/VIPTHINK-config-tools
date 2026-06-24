from __future__ import annotations

import json
import os
import subprocess
import sys
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
CONFIG_FILE = ROOT / "简案配置表.xlsx"
CAPTURE_DIR = ROOT / "简案接口监听"
HOST = "127.0.0.1"
PORT = int(os.environ.get("VIPTHINK_JIANAN_ASSISTANT_PORT", "8771"))
SHEET_NAME = "简案自动配置"

HEADERS = [
    "源课类ID",
    "源课类名称",
    "源课件代码",
    "目标课类ID",
    "目标课类名称",
    "目标课件代码",
    "简案文件名",
    "下载状态",
    "上传状态",
    "备注",
]

SAMPLE_ROW = [
    "3203,3278",
    "示例课类A",
    "s4_v8_01_YY",
    "3203,3278",
    "示例课类B",
    "s4_v8_01_TW",
    "",
    "待下载",
    "待上传",
    "一行表示一组源到目标的映射关系",
]


def text(value) -> str:
    return "" if value is None else str(value).strip()


def json_bytes(payload: object) -> bytes:
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def ensure_config_file() -> None:
    if CONFIG_FILE.exists():
        workbook = openpyxl.load_workbook(CONFIG_FILE)
        sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
        sheet.title = SHEET_NAME
        for index, header in enumerate(HEADERS, start=1):
            if text(sheet.cell(1, index).value) != header:
                sheet.cell(1, index).value = header
        workbook.save(CONFIG_FILE)
        return

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(HEADERS)
    sheet.append(SAMPLE_ROW)
    for col in range(1, len(HEADERS) + 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    workbook.save(CONFIG_FILE)


def load_tasks() -> list[dict]:
    ensure_config_file()
    workbook = openpyxl.load_workbook(CONFIG_FILE, data_only=True)
    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
    tasks = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        source_code = text(row[2] if len(row) > 2 else "")
        target_code = text(row[5] if len(row) > 5 else "")
        if not source_code and not target_code:
          continue
        tasks.append({
            "row": row_number,
            "source_course_category_id": text(row[0] if len(row) > 0 else ""),
            "source_course_category_name": text(row[1] if len(row) > 1 else ""),
            "source_code": source_code,
            "target_course_category_id": text(row[3] if len(row) > 3 else ""),
            "target_course_category_name": text(row[4] if len(row) > 4 else ""),
            "target_code": target_code,
            "filename": text(row[6] if len(row) > 6 else ""),
            "download_status": text(row[7] if len(row) > 7 else ""),
            "upload_status": text(row[8] if len(row) > 8 else ""),
            "note": text(row[9] if len(row) > 9 else ""),
        })
    return tasks


def save_task_results(results: list[dict]) -> None:
    ensure_config_file()
    workbook = openpyxl.load_workbook(CONFIG_FILE)
    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
    for result in results:
        row = int(result.get("row") or 0)
        if row < 2:
            continue
        download_status = text(result.get("download_status"))
        upload_status = text(result.get("upload_status"))
        note = text(result.get("note"))
        if download_status:
            sheet.cell(row, 8).value = download_status
        if upload_status:
            sheet.cell(row, 9).value = upload_status
        sheet.cell(row, 10).value = note
    workbook.save(CONFIG_FILE)


def build_plan() -> dict:
    tasks = load_tasks()
    ready = [
        task for task in tasks
        if task["source_code"] and task["target_code"]
    ]
    missing = [task for task in tasks if task not in ready]
    return {
        "ok": True,
        "count": len(tasks),
        "ready_count": len(ready),
        "missing_count": len(missing),
        "tasks": tasks,
        "message": "当前版本已建立数据流，执行前会先输出任务计划。下载/上传真实接口待监听结果补齐。",
    }


def run_capture_script() -> dict:
    script = ROOT / "listen_jianan_requests.py"
    if not script.exists():
        raise FileNotFoundError(f"找不到监听脚本：{script}")
    creation = subprocess.Popen(
        [sys.executable, str(script)],
        cwd=str(ROOT),
        creationflags=getattr(subprocess, "CREATE_NEW_CONSOLE", 0),
    )
    return {
        "ok": True,
        "message": "已启动监听脚本，请在新浏览器里完成一次简案下载和上传操作。",
        "pid": creation.pid,
        "capture_dir": str(CAPTURE_DIR),
    }


def config_help() -> dict:
    ensure_config_file()
    return {
        "ok": True,
        "config_file": str(CONFIG_FILE),
        "headers": HEADERS,
        "tips": [
            "源课类ID、目标课类ID建议填写页面 URL 中的 course_category_id，例如 3203,3278。",
            "一行维护一组源课件代码到目标课件代码的对应关系。",
            "简案文件名可先留空，后续监听到真实下载命名规则后可自动回填。",
        ],
    }


def capture_guide() -> dict:
    return {
        "ok": True,
        "url": "https://jy.vipthink.cn/#/teacher_new/plan?course_category_id=3203%2C3278&game_url=&order=asc&page=1&page_count=10&page_num=1&sort=cn.sort&subject=-1",
        "steps": [
            "启动监听脚本。",
            "在新打开的浏览器里登录 jy.vipthink.cn。",
            "进入简案列表页面，手动下载 1 个简案 Excel。",
            "再手动把该简案上传到目标课件。",
            "把监听结果文件留在“简案接口监听”目录，后续可直接补代码。",
        ],
    }


class ReusableThreadingHTTPServer(ThreadingHTTPServer):
    allow_reuse_address = True
    daemon_threads = True


class Handler(BaseHTTPRequestHandler):
    server_version = "VipthinkJiananAssistant/0.1"

    def end_headers(self) -> None:
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        super().end_headers()

    def do_OPTIONS(self) -> None:
        self.send_response(204)
        self.end_headers()

    def send_json(self, payload: object, status: int = 200) -> None:
        data = json_bytes(payload)
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self) -> None:
        try:
            if self.path == "/health":
                ensure_config_file()
                self.send_json({
                    "ok": True,
                    "port": PORT,
                    "root": str(ROOT),
                    "config_file": str(CONFIG_FILE),
                    "capture_dir": str(CAPTURE_DIR),
                })
                return
            if self.path == "/config-help":
                self.send_json(config_help())
                return
            if self.path == "/capture-guide":
                self.send_json(capture_guide())
                return
            self.send_json({"ok": False, "error": "Not found"}, 404)
        except Exception as error:
            self.send_json({"ok": False, "error": str(error)}, 500)

    def do_POST(self) -> None:
        length = int(self.headers.get("Content-Length") or 0)
        body = self.rfile.read(length).decode("utf-8") if length else "{}"
        try:
            payload = json.loads(body)
        except json.JSONDecodeError:
            payload = {}

        try:
            if self.path == "/prepare-workbook":
                ensure_config_file()
                self.send_json({"ok": True, "config_file": str(CONFIG_FILE)})
                return
            if self.path == "/run-capture":
                self.send_json(run_capture_script())
                return
            if self.path == "/report":
                results = payload.get("results") or []
                save_task_results(results)
                self.send_json({"ok": True, "count": len(results)})
                return
            if self.path == "/execute":
                plan = build_plan()
                dry_run = bool(payload.get("dryRun"))
                session_id = text(payload.get("sessionId"))
                plan["dry_run"] = dry_run
                plan["session_id_captured"] = bool(session_id)
                if not session_id:
                    plan["ok"] = False
                    plan["error"] = "未收到 Session-Id，请先在页面中刷新一次。"
                if not dry_run and plan["ok"]:
                    plan["message"] = "本地助手已返回任务列表，可由插件直接按源课件简案复制到目标课件。"
                self.send_json(plan)
                return
            self.send_json({"ok": False, "error": "Not found"}, 404)
        except Exception as error:
            self.send_json({"ok": False, "error": str(error)}, 500)

    def log_message(self, fmt: str, *args) -> None:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {fmt % args}")


def main() -> None:
    ensure_config_file()
    CAPTURE_DIR.mkdir(exist_ok=True)
    print(f"简案自动配置本地助手已启动：http://{HOST}:{PORT}")
    print(f"配置表：{CONFIG_FILE}")
    print("请保持窗口打开，然后在插件中执行检查或启动监听。")
    server = ReusableThreadingHTTPServer((HOST, PORT), Handler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()
        print("简案自动配置本地助手已关闭。")


if __name__ == "__main__":
    main()
