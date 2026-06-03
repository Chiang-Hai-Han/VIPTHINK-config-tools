from __future__ import annotations

import json
import os
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import openpyxl


ROOT = Path(__file__).resolve().parent
CONFIG_FILE = ROOT / "讲次配置表.xlsx"
CHAPTER_CONFIG_DIR = ROOT / "讲次配置结果"
HOST = "127.0.0.1"
PORT = int(os.environ.get("VIPTHINK_CHAPTER_CONFIG_PORT", "8768"))
SHEET_NAME = "讲次配置"

HEADERS = [
    "讲次ID(cn_id)",
    "课类",
    "关联课件",
    "讲次名称",
    "故事场景",
    "上课人数",
    "粤语讲次代码",
    "粤语课类",
    "配置状态",
    "备注",
]


def text(value) -> str:
    return "" if value is None else str(value).strip()


def json_bytes(payload: object) -> bytes:
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def ensure_config_sheet():
    if CONFIG_FILE.exists():
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    for col, header in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col, value=header)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 30
    wb.save(str(CONFIG_FILE))
    print(f"已创建讲次配置表：{CONFIG_FILE}")


def read_tasks():
    if not CONFIG_FILE.exists():
        raise FileNotFoundError(f"讲次配置表不存在：{CONFIG_FILE}")
    wb = openpyxl.load_workbook(str(CONFIG_FILE))
    ws = wb.active
    tasks = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        cells = [c.value for c in row]
        if not any(cells):
            break
        code = text(cells[0])
        if not code:
            continue
        tasks.append({
            "row": row[0].row,
            "讲次ID(cn_id)": code,
            "课类": text(cells[1] if len(cells) > 1 else ""),
            "关联课件": text(cells[2] if len(cells) > 2 else ""),
            "讲次名称": text(cells[3] if len(cells) > 3 else ""),
            "故事场景": text(cells[4] if len(cells) > 4 else ""),
            "上课人数": text(cells[5] if len(cells) > 5 else ""),
            "粤语讲次代码": text(cells[6] if len(cells) > 6 else ""),
            "粤语课类": text(cells[7] if len(cells) > 7 else ""),
            "配置状态": text(cells[8] if len(cells) > 8 else ""),
            "备注": text(cells[9] if len(cells) > 9 else ""),
        })
    return tasks


def save_report(tasks):
    if not CONFIG_FILE.exists():
        return
    wb = openpyxl.load_workbook(str(CONFIG_FILE))
    ws = wb.active
    CHAPTER_CONFIG_DIR.mkdir(exist_ok=True)
    results = []
    for task in tasks:
        row_num = task.get("row")
        if row_num:
            ws.cell(row=row_num, column=9, value=task.get("配置状态", ""))
            ws.cell(row=row_num, column=10, value=task.get("备注", ""))
        results.append({
            "讲次ID(cn_id)": task.get("讲次ID(cn_id)", ""),
            "课类": task.get("课类", ""),
            "关联课件": task.get("关联课件", ""),
            "讲次名称": task.get("讲次名称", ""),
            "故事场景": task.get("故事场景", ""),
            "上课人数": task.get("上课人数", ""),
            "粤语讲次代码": task.get("粤语讲次代码",
    "粤语课类", ""),
            "配置状态": task.get("配置状态", ""),
            "讲次ID": task.get("讲次ID", ""),
            "课件ID": task.get("课件ID", ""),
            "课件名称": task.get("课件名称", ""),
            "备注": task.get("备注", ""),
            "错误信息": task.get("错误信息", ""),
        })
    wb.save(str(CONFIG_FILE))
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    result_file = CHAPTER_CONFIG_DIR / f"讲次配置结果_{timestamp}.json"
    result_file.write_text(
        json.dumps(results, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )


class Handler(BaseHTTPRequestHandler):

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if path == "/health":
            self.send_json({"ok": True, "port": PORT, "config": str(CONFIG_FILE)})
        elif path == "/tasks":
            try:
                tasks = read_tasks()
                self.send_json({"ok": True, "tasks": tasks})
            except Exception as error:
                self.send_json({"ok": False, "error": str(error)}, status=500)
        else:
            self.send_json({"ok": False, "error": f"未知路径：{path}"}, status=404)

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path
        length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(length) if length else b"{}"
        data = json.loads(body.decode("utf-8")) if body else {}
        if path == "/report":
            tasks = data.get("tasks", [])
            try:
                save_report(tasks)
                self.send_json({"ok": True})
            except Exception as error:
                self.send_json({"ok": False, "error": str(error)}, status=500)
        else:
            self.send_json({"ok": False, "error": f"未知路径：{path}"}, status=404)

    def send_json(self, payload: object, status: int = 200):
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(json_bytes(payload))

    def log_message(self, fmt, *args):
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {fmt % args}")


def main():
    ensure_config_sheet()
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    print(f"讲次配置助手已启动：http://{HOST}:{PORT}")
    print(f"配置文件：{CONFIG_FILE}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n已关闭。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

