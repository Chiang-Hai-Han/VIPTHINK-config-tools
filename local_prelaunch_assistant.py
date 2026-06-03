from __future__ import annotations

import json
import mimetypes
import os
import re
import shutil
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
from copy import copy
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import openpyxl


ROOT = Path(__file__).resolve().parent
CONFIG_FILE = ROOT / "课件上架前配置表.xlsx"
POSTLAUNCH_CONFIG_FILE = ROOT / "课件上架后配置表.xlsx"
SMALL_TEACHER_CONFIG_FILE = ROOT / "小老师配置表.xlsx"
RESULT_FILE = ROOT / "课件上架前配置结果.json"
IMAGE_DIR = ROOT / "待上传图片文件夹"
SMALL_TEACHER_IMAGE_DIR = ROOT / "待上传小老师图片文件夹"
TEMPLATE_FILE = ROOT / "resource-copy-template.json"
COURSEWARE_DATA_DIR = ROOT / "课件数据"
PENDING_DATA_DIR = ROOT / "待处理数据"
IMPORT_TEMPLATE_FILE = Path(r"C:\Users\jianghaihan\Documents\自动下载图片\批量新增课件模板 (1).xlsx")
SHEET_NAME = "课件上架前配置"
HOST = "127.0.0.1"
PORT = int(os.environ.get("VIPTHINK_PRELAUNCH_ASSISTANT_PORT", "8769"))
IMAGE_SUFFIXES = [".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"]
IMPORT_HEADER_ROW = 2
IMPORT_DATA_START_ROW = 3

HEADERS = [
    "课件编码",
    "课件名称",
    "封面图片",
    "作业课件编码",
    "资源来源课件",
    "封面上传状态",
    "课件关联状态",
    "资源复制状态",
    "备注",
]

POSTLAUNCH_HEADERS = [
    "课件编码",
    "资源来源课件",
    "知识点复制状态",
    "小老师上传状态",
    "备注",
]

SMALL_TEACHER_CONFIG_HEADERS = [
    "课件编码",
    "建议主题",
    "录制建议",
    "图片资源",
    "上传状态",
    "备注",
]

COURSEWARE_DATA_HEADERS = [
    "序号",
    "列表_课件名称",
    "列表_课件代码",
    "列表_课程封面",
    "详情_课件名称",
    "详情_内容名称",
    "详情_课件编码",
    "详情_教学目标",
    "详情_课节简介",
]

SMALL_TEACHER_HEADERS = [
    "序号",
    "课件编码",
    "建议主题",
    "录制建议",
    "图片资源",
    "本地图片",
]


def json_bytes(payload: object) -> bytes:
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def text(value) -> str:
    return "" if value is None else str(value).strip()


def safe_filename(value: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|]+', "_", text(value))
    cleaned = re.sub(r"\s+", "_", cleaned).strip("._ ")
    return cleaned or "课件信息"


def normalize_paragraphs(value: str) -> str:
    raw = text(value)
    if not raw:
        return ""
    if re.fullmatch(r"\s*(?:<p\b[^>]*>.*?</p>\s*)+\s*", raw, flags=re.IGNORECASE | re.DOTALL):
        return raw
    if not re.search(r"\r\n|\r|\n", raw):
        return raw
    parts = [part.strip() for part in re.split(r"\r\n|\r|\n", raw)]
    return "".join(f"<p>{part}</p>" for part in parts if part)


def write_courseware_data(payload: dict) -> Path:
    search_name = text(payload.get("searchName") or payload.get("search_name"))
    rows = payload.get("rows") or []
    if not isinstance(rows, list):
        raise RuntimeError("课件信息数据格式不正确：rows 必须是列表。")

    COURSEWARE_DATA_DIR.mkdir(exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "课件信息"
    sheet.append(COURSEWARE_DATA_HEADERS)

    for index, item in enumerate(rows, start=1):
        item = item if isinstance(item, dict) else {}
        sheet.append([
            index,
            text(item.get("列表_课件名称")),
            text(item.get("列表_课件代码")),
            text(item.get("列表_课程封面")),
            text(item.get("详情_课件名称")),
            text(item.get("详情_内容名称")),
            text(item.get("详情_课件编码")),
            text(item.get("详情_教学目标")),
            text(item.get("详情_课节简介")),
        ])

    for column_cells in sheet.columns:
        max_length = max(len(text(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 50)

    filename = f"{safe_filename(search_name)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = COURSEWARE_DATA_DIR / filename
    workbook.save(output_path)
    return output_path


def write_small_teacher_data(payload: dict) -> dict:
    search_name = text(payload.get("searchName") or payload.get("search_name"))
    rows = payload.get("rows") or []
    if not isinstance(rows, list):
        raise RuntimeError("小老师数据格式不正确：rows 必须是列表。")

    COURSEWARE_DATA_DIR.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = COURSEWARE_DATA_DIR / f"小老师{safe_filename(search_name)}_{stamp}.xlsx"
    image_dir = COURSEWARE_DATA_DIR / f"小老师{safe_filename(search_name)}_{stamp}"
    image_dir.mkdir(exist_ok=True)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "小老师"
    sheet.append(SMALL_TEACHER_HEADERS)

    downloaded = 0
    failed: list[dict[str, str]] = []
    for index, item in enumerate(rows, start=1):
        item = item if isinstance(item, dict) else {}
        code = text(item.get("课件编码"))
        image_url = text(item.get("图片资源"))
        local_image = ""
        if code and image_url:
            try:
                local_image = str(download_cover(image_url, image_dir, f"小老师{code}"))
                downloaded += 1
            except Exception as error:
                failed.append({"code": code, "error": str(error)})
        sheet.append([
            index,
            code,
            text(item.get("建议主题")),
            text(item.get("录制建议")),
            image_url,
            local_image,
        ])

    for column_cells in sheet.columns:
        max_length = max(len(text(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 50)

    workbook.save(output_path)
    return {
        "path": str(output_path),
        "image_dir": str(image_dir),
        "count": len(rows),
        "downloaded": downloaded,
        "failed": failed,
    }


def latest_courseware_data_file() -> Path:
    files = [
        path for path in COURSEWARE_DATA_DIR.glob("*.xlsx")
        if "_已填入模板" not in path.stem and not path.name.startswith("~$")
    ]
    if not files:
        raise FileNotFoundError(f"课件数据文件夹中没有可填表的 Excel：{COURSEWARE_DATA_DIR}")
    return max(files, key=lambda path: path.stat().st_mtime)


def read_workbook_rows(path: Path) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.active
    headers = [text(cell.value) for cell in sheet[1]]
    rows: list[dict[str, str]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = {headers[index]: text(value) for index, value in enumerate(values) if index < len(headers)}
        if any(row.values()):
            rows.append(row)
    return rows


def workbook_headers(sheet, row_number: int) -> dict[str, int]:
    return {
        text(sheet.cell(row_number, column).value): column
        for column in range(1, sheet.max_column + 1)
        if text(sheet.cell(row_number, column).value)
    }


def first_row_value(row: dict[str, str], candidates: list[str]) -> str:
    for candidate in candidates:
        value = text(row.get(candidate))
        if value:
            return value
    return ""


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


def template_options() -> dict[str, list[str]]:
    if not IMPORT_TEMPLATE_FILE.exists():
        raise FileNotFoundError(f"找不到模板：{IMPORT_TEMPLATE_FILE}")
    workbook = openpyxl.load_workbook(IMPORT_TEMPLATE_FILE, data_only=True)
    if len(workbook.worksheets) < 2:
        raise RuntimeError("模板缺少下拉选项工作表。")
    sheet = workbook.worksheets[1]
    headers = workbook_headers(sheet, 1)
    wanted = ["语种", "科目", "课件类型", "课件难度", "课件版本"]
    result: dict[str, list[str]] = {}
    for name in wanted:
        column = headers.get(name)
        values: list[str] = []
        if column:
            for row in range(2, sheet.max_row + 1):
                value = text(sheet.cell(row, column).value)
                if value:
                    values.append(value)
        result[name] = values
    return result


def normalize_url(url: str) -> str:
    url = text(url)
    if url.startswith("//"):
        return "https:" + url
    if url and not urllib.parse.urlparse(url).scheme:
        return "https://" + url
    return url


def image_extension(url: str, response) -> str:
    suffix = Path(urllib.parse.unquote(urllib.parse.urlparse(url).path)).suffix
    if suffix and len(suffix) <= 8:
        return ".jpg" if suffix.lower() == ".jpeg" else suffix
    content_type = response.headers.get("Content-Type", "").split(";")[0].strip().lower()
    guessed = mimetypes.guess_extension(content_type)
    return ".jpg" if guessed in {"", None, ".jpe"} else guessed


def download_cover(url: str, output_dir: Path, code: str, timeout: int = 25, retries: int = 2) -> Path:
    os.environ.pop("SSLKEYLOGFILE", None)
    request = urllib.request.Request(
        normalize_url(url),
        headers={
            "User-Agent": "Mozilla/5.0 course-cover-downloader",
            "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
        },
    )
    last_error: Exception | None = None
    for attempt in range(retries + 1):
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                suffix = image_extension(url, response)
                output_path = output_dir / f"{safe_filename(code)}{suffix}"
                output_path.write_bytes(response.read())
                return output_path
        except (urllib.error.URLError, TimeoutError, OSError) as exc:
            last_error = exc
            if attempt < retries:
                time.sleep(1.5)
    raise RuntimeError(str(last_error) if last_error else "下载失败")


def fill_latest_data_to_template(options: dict) -> dict:
    source_path = latest_courseware_data_file()
    rows = read_workbook_rows(source_path)
    if not rows:
        raise RuntimeError(f"最近一次课件数据为空：{source_path}")
    if not IMPORT_TEMPLATE_FILE.exists():
        raise FileNotFoundError(f"找不到模板：{IMPORT_TEMPLATE_FILE}")

    PENDING_DATA_DIR.mkdir(exist_ok=True)
    batch_name = f"{source_path.stem}_已填入模板"
    batch_dir = PENDING_DATA_DIR / batch_name
    batch_dir.mkdir(exist_ok=True)
    output_path = batch_dir / f"{batch_name}.xlsx"
    shutil.copy2(IMPORT_TEMPLATE_FILE, output_path)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.worksheets[0]
    headers = workbook_headers(sheet, IMPORT_HEADER_ROW)
    style_row = IMPORT_DATA_START_ROW

    selected = {
        "*语种": text(options.get("语种")),
        "*科目": text(options.get("科目")),
        "*课件类型": text(options.get("课件类型")),
        "*课件难度": text(options.get("课件难度")),
        "*课件版本": text(options.get("课件版本")),
    }
    row_mappings = {
        "*课件名称": ["详情_课件名称", "列表_课件名称"],
        "内容名称": ["详情_内容名称"],
        "*课件编码": ["列表_课件代码", "详情_课件编码"],
        "教学目标": ["详情_教学目标"],
        "课节简介": ["详情_课节简介"],
    }

    max_rows = max(sheet.max_row, IMPORT_DATA_START_ROW + len(rows) - 1)
    for excel_row in range(IMPORT_DATA_START_ROW, max_rows + 1):
        for column in range(1, sheet.max_column + 1):
            copy_cell_style(sheet.cell(style_row, column), sheet.cell(excel_row, column))
            sheet.cell(excel_row, column).value = None

    for index, source_row in enumerate(rows, start=IMPORT_DATA_START_ROW):
        for template_header, candidates in row_mappings.items():
            column = headers.get(template_header)
            if not column:
                continue
            value = first_row_value(source_row, candidates)
            if template_header in {"教学目标", "课节简介"}:
                value = normalize_paragraphs(value)
            sheet.cell(index, column).value = value

        for template_header, value in selected.items():
            column = headers.get(template_header)
            if column:
                sheet.cell(index, column).value = value

    workbook.save(output_path)

    image_dir = batch_dir / "课程封面"
    image_dir.mkdir(exist_ok=True)
    downloaded = 0
    failed: list[dict[str, str]] = []
    for source_row in rows:
        code = first_row_value(source_row, ["列表_课件代码", "详情_课件编码", "课件编码"])
        cover_url = first_row_value(source_row, ["列表_课程封面", "课程封面"])
        if not code or not cover_url:
            continue
        try:
            download_cover(cover_url, image_dir, code)
            downloaded += 1
        except Exception as error:
            failed.append({"code": code, "error": str(error)})

    return {
        "source": str(source_path),
        "batch_dir": str(batch_dir),
        "output": str(output_path),
        "image_dir": str(image_dir),
        "rows": len(rows),
        "downloaded": downloaded,
        "failed": failed,
    }


def ensure_config_file() -> None:
    if CONFIG_FILE.exists():
        workbook = openpyxl.load_workbook(CONFIG_FILE)
        sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
        sheet.title = SHEET_NAME
        existing = [text(cell.value) for cell in sheet[1]]
        changed = False
        for col, header in enumerate(HEADERS, start=1):
            if col > len(existing) or existing[col - 1] != header:
                sheet.cell(1, col).value = header
                changed = True
        if changed:
            try:
                workbook.save(CONFIG_FILE)
            except PermissionError:
                print(f"配置表正在被打开，已跳过自动保存：{CONFIG_FILE}")
        return

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(HEADERS)
    sheet.append([
        "s4_v8_04_TW",
        "",
        "",
        "s4_v8_04_TW_hw",
        "s4_v8_04_YY",
        "待上传",
        "待关联",
        "待复制",
        "示例：同一行维护封面、线上作业关联、资源复制的编码对应关系",
    ])
    workbook.save(CONFIG_FILE)


def ensure_simple_sheet(path: Path, sheet_name: str, headers: list[str], sample_row: list[str]) -> None:
    if path.exists():
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        sheet.title = sheet_name
        existing = [text(cell.value) for cell in sheet[1]]
        changed = False
        for col, header in enumerate(headers, start=1):
            if col > len(existing) or existing[col - 1] != header:
                sheet.cell(1, col).value = header
                changed = True
        if changed:
            workbook.save(path)
        return
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    sheet.append(sample_row)
    workbook.save(path)


def ensure_postlaunch_config_file() -> None:
    ensure_simple_sheet(
        POSTLAUNCH_CONFIG_FILE,
        "课件上架后配置",
        POSTLAUNCH_HEADERS,
        ["s4_v8_01_TW", "s4_v8_01_YY", "待复制", "待上传", "测试集"],
    )


def ensure_small_teacher_config_file() -> None:
    ensure_simple_sheet(
        SMALL_TEACHER_CONFIG_FILE,
        "小老师配置",
        SMALL_TEACHER_CONFIG_HEADERS,
        ["s4_v8_01_TW", "", "", "", "待上传", "测试集"],
    )


def workbook_candidates(kind: str) -> list[dict]:
    patterns = {
        "prelaunch": "*上架前配置表*.xlsx",
        "postlaunch": "*上架后配置表*.xlsx",
        "small_teacher": "*小老师配置表*.xlsx",
    }
    files = [
        path for path in ROOT.glob(patterns[kind])
        if path.is_file() and not path.name.startswith("~$")
    ]
    return [{"name": path.name, "path": str(path)} for path in sorted(files, key=lambda item: item.stat().st_mtime, reverse=True)]


def resolve_workbook_path(kind: str, selected_path: str = "") -> Path:
    if selected_path:
        path = Path(selected_path)
        if not path.is_absolute():
            path = ROOT / path
        if not path.exists():
            raise FileNotFoundError(f"找不到选择的表格：{path}")
        return path
    defaults = {
        "prelaunch": CONFIG_FILE,
        "postlaunch": POSTLAUNCH_CONFIG_FILE,
        "small_teacher": SMALL_TEACHER_CONFIG_FILE,
    }
    return defaults[kind]


def header_index(headers: list[str], candidates: list[str]) -> int | None:
    normalized = {header.lower(): index for index, header in enumerate(headers)}
    for candidate in candidates:
        index = normalized.get(candidate.lower())
        if index is not None:
            return index
    return None


def load_rows(config_path: Path | None = None) -> list[dict]:
    if config_path is None:
        ensure_config_file()
        config_path = CONFIG_FILE
    workbook = openpyxl.load_workbook(config_path, data_only=True)
    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
    headers = [text(cell.value) for cell in sheet[1]]
    idx = {
        "code": header_index(headers, ["课件编码", "courseware_code", "code"]),
        "name": header_index(headers, ["课件名称", "courseware_name", "name"]),
        "image": header_index(headers, ["封面图片", "image", "image_file"]),
        "homework": header_index(headers, ["作业课件编码", "线上作业编码", "homework_code", "target_code"]),
        "resource_source": header_index(headers, ["资源来源课件", "source_code", "copy_from_code"]),
        "cover_status": header_index(headers, ["封面上传状态"]),
        "relation_status": header_index(headers, ["课件关联状态", "线上作业关联状态"]),
        "copy_status": header_index(headers, ["资源复制状态"]),
        "note": header_index(headers, ["备注", "note"]),
    }
    if idx["code"] is None:
        raise RuntimeError(f"{config_path.name} 缺少列：课件编码")

    rows: list[dict] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        code = text(row[idx["code"]] if idx["code"] is not None and idx["code"] < len(row) else "")
        if not code:
            continue
        homework = text(row[idx["homework"]] if idx["homework"] is not None and idx["homework"] < len(row) else "")
        resource_source = text(row[idx["resource_source"]] if idx["resource_source"] is not None and idx["resource_source"] < len(row) else "")
        rows.append({
            "row": row_number,
            "courseware_code": code,
            "courseware_name": text(row[idx["name"]] if idx["name"] is not None and idx["name"] < len(row) else ""),
            "image_file_in_sheet": text(row[idx["image"]] if idx["image"] is not None and idx["image"] < len(row) else ""),
            "homework_code": homework or f"{code}_hw",
            "resource_source_code": resource_source or code.replace("_TW", "_YY"),
            "cover_status": text(row[idx["cover_status"]] if idx["cover_status"] is not None and idx["cover_status"] < len(row) else ""),
            "relation_status": text(row[idx["relation_status"]] if idx["relation_status"] is not None and idx["relation_status"] < len(row) else ""),
            "copy_status": text(row[idx["copy_status"]] if idx["copy_status"] is not None and idx["copy_status"] < len(row) else ""),
            "note": text(row[idx["note"]] if idx["note"] is not None and idx["note"] < len(row) else ""),
        })
    return rows


def find_image(filename: str, code: str) -> Path | None:
    raw = text(filename)
    candidates: list[Path] = []
    if raw:
        raw_path = Path(raw)
        if raw_path.suffix:
            candidates.append(IMAGE_DIR / raw_path.name)
            candidates.append(raw_path)
        else:
            candidates.extend(IMAGE_DIR / f"{raw_path.name}{suffix}" for suffix in IMAGE_SUFFIXES)
    if code:
        candidates.extend(IMAGE_DIR / f"{code}{suffix}" for suffix in IMAGE_SUFFIXES)

    seen: set[Path] = set()
    for candidate in candidates:
        normalized = candidate.resolve()
        if normalized in seen:
            continue
        seen.add(normalized)
        if candidate.exists() and candidate.is_file():
            return candidate
    return None


def find_small_teacher_image(filename: str, code: str = "") -> Path | None:
    raw = text(filename)
    candidates: list[Path] = []
    if raw:
        raw_path = Path(raw)
        if raw_path.is_absolute():
            candidates.append(raw_path)
        elif raw_path.suffix:
            candidates.append(SMALL_TEACHER_IMAGE_DIR / raw_path.name)
            candidates.append(ROOT / raw_path)
            candidates.append(IMAGE_DIR / raw_path.name)
        else:
            candidates.extend(SMALL_TEACHER_IMAGE_DIR / f"{raw_path}{suffix}" for suffix in IMAGE_SUFFIXES)
            candidates.extend(ROOT / f"{raw_path}{suffix}" for suffix in IMAGE_SUFFIXES)
            candidates.extend(IMAGE_DIR / f"{raw_path}{suffix}" for suffix in IMAGE_SUFFIXES)
    if code:
        base = re.sub(r"_(?:TW|YY)$", "_", code, flags=re.IGNORECASE)
        if base and base != code:
            for path in SMALL_TEACHER_IMAGE_DIR.glob(f"{safe_filename(base)}*"):
                if path.suffix.lower() in IMAGE_SUFFIXES:
                    candidates.append(path)
        candidates.extend(ROOT / f"小老师{code}{suffix}" for suffix in IMAGE_SUFFIXES)
        candidates.extend(SMALL_TEACHER_IMAGE_DIR / f"小老师{code}{suffix}" for suffix in IMAGE_SUFFIXES)
        candidates.extend(SMALL_TEACHER_IMAGE_DIR / f"{code}{suffix}" for suffix in IMAGE_SUFFIXES)
        candidates.extend(IMAGE_DIR / f"小老师{code}{suffix}" for suffix in IMAGE_SUFFIXES)
        candidates.extend(IMAGE_DIR / f"{code}{suffix}" for suffix in IMAGE_SUFFIXES)

    seen: set[Path] = set()
    for candidate in candidates:
        try:
            normalized = candidate.resolve()
        except OSError:
            continue
        if normalized in seen:
            continue
        seen.add(normalized)
        if normalized.exists() and normalized.is_file():
            return normalized
    return None


def build_cover_tasks(config_path: Path | None = None) -> list[dict]:
    tasks = []
    for row in load_rows(config_path):
        image_path = find_image(row["image_file_in_sheet"], row["courseware_code"])
        tasks.append({
            "row": row["row"],
            "courseware_code": row["courseware_code"],
            "courseware_name": row["courseware_name"],
            "image_file_in_sheet": row["image_file_in_sheet"],
            "image_path": str(image_path.resolve()) if image_path else "",
            "image_exists": image_path is not None,
            "status": row["cover_status"],
            "completed": row["cover_status"] in {"已上传", "跳过"},
            "note": row["note"],
        })
    return tasks


def build_relation_tasks(config_path: Path | None = None) -> list[dict]:
    return [
        {
            "row": row["row"],
            "source_code": row["courseware_code"],
            "target_code": row["homework_code"],
            "status": row["relation_status"],
            "completed": row["relation_status"] in {"已关联", "跳过"},
            "note": row["note"],
        }
        for row in load_rows(config_path)
    ]


def build_resource_copy_tasks(config_path: Path | None = None) -> list[dict]:
    return [
        {
            "row": row["row"],
            "target_code": row["courseware_code"],
            "source_code": row["resource_source_code"],
            "status": row["copy_status"],
            "completed": row["copy_status"] in {"已复制", "跳过"},
            "note": row["note"],
        }
        for row in load_rows(config_path)
    ]


def load_postlaunch_rows(config_path: Path | None = None) -> list[dict]:
    ensure_postlaunch_config_file()
    config_path = config_path or POSTLAUNCH_CONFIG_FILE
    workbook = openpyxl.load_workbook(config_path, data_only=True)
    sheet = workbook["课件上架后配置"] if "课件上架后配置" in workbook.sheetnames else workbook.active
    headers = [text(cell.value) for cell in sheet[1]]
    code_idx = header_index(headers, ["课件编码", "target_code", "code"])
    source_idx = header_index(headers, ["资源来源课件", "source_code", "copy_from_code"])
    status_idx = header_index(headers, ["知识点复制状态"])
    note_idx = header_index(headers, ["备注", "note"])
    if code_idx is None or source_idx is None:
        raise RuntimeError(f"{config_path.name} 缺少列：课件编码 或 资源来源课件")
    rows = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        code = text(row[code_idx] if code_idx < len(row) else "")
        source = text(row[source_idx] if source_idx < len(row) else "")
        if not code or not source:
            continue
        rows.append({
            "row": row_number,
            "target_code": code,
            "source_code": source,
            "status": text(row[status_idx] if status_idx is not None and status_idx < len(row) else ""),
            "completed": text(row[status_idx] if status_idx is not None and status_idx < len(row) else "") in {"已复制", "跳过"},
            "note": text(row[note_idx] if note_idx is not None and note_idx < len(row) else ""),
        })
    return rows


def build_knowledge_copy_tasks(config_path: Path | None = None) -> list[dict]:
    return load_postlaunch_rows(config_path)


def build_small_teacher_upload_tasks(config_path: Path | None = None) -> list[dict]:
    ensure_small_teacher_config_file()
    config_path = config_path or SMALL_TEACHER_CONFIG_FILE
    workbook = openpyxl.load_workbook(config_path, data_only=True)
    sheet = workbook["小老师配置"] if "小老师配置" in workbook.sheetnames else workbook.active
    headers = [text(cell.value) for cell in sheet[1]]
    idx = {
        "code": header_index(headers, ["课件编码", "code"]),
        "name": header_index(headers, ["建议主题", "name"]),
        "advice": header_index(headers, ["录制建议", "advice"]),
        "image": header_index(headers, ["图片资源", "image", "image_file"]),
        "status": header_index(headers, ["上传状态"]),
        "note": header_index(headers, ["备注", "note"]),
    }
    if idx["code"] is None:
        raise RuntimeError(f"{config_path.name} 缺少列：课件编码")
    tasks = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        code = text(row[idx["code"]] if idx["code"] < len(row) else "")
        if not code:
            continue
        tasks.append({
            "row": row_number,
            "target_code": code,
            "name": text(row[idx["name"]] if idx["name"] is not None and idx["name"] < len(row) else ""),
            "advice": text(row[idx["advice"]] if idx["advice"] is not None and idx["advice"] < len(row) else ""),
            "image": text(row[idx["image"]] if idx["image"] is not None and idx["image"] < len(row) else ""),
            "status": text(row[idx["status"]] if idx["status"] is not None and idx["status"] < len(row) else ""),
            "completed": text(row[idx["status"]] if idx["status"] is not None and idx["status"] < len(row) else "") in {"已上传", "跳过"},
            "note": text(row[idx["note"]] if idx["note"] is not None and idx["note"] < len(row) else ""),
        })
    return tasks


def small_teacher_task_by_row_or_code(row: str = "", code: str = "") -> dict | None:
    tasks = build_small_teacher_upload_tasks()
    for task in tasks:
        if row and str(task.get("row")) == str(row):
            return task
        if code and task.get("target_code") == code:
            return task
    return None


def load_resource_copy_template() -> dict:
    if TEMPLATE_FILE.exists():
        return json.loads(TEMPLATE_FILE.read_text(encoding="utf-8"))
    raise RuntimeError("缺少 resource-copy-template.json，请重新监听一次资源复制操作。")


def append_result(kind: str, payload: dict) -> None:
    records = []
    if RESULT_FILE.exists():
        try:
            records = json.loads(RESULT_FILE.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            records = []
    records.append({"kind": kind, **payload})
    RESULT_FILE.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")


def update_sheet_result(kind: str, payload: dict) -> None:
    ensure_config_file()
    workbook = openpyxl.load_workbook(CONFIG_FILE)
    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
    headers = [text(cell.value) for cell in sheet[1]]
    status_header = {
        "cover": "封面上传状态",
        "relation": "课件关联状态",
        "resource_copy": "资源复制状态",
    }[kind]
    status_idx = header_index(headers, [status_header])
    note_idx = header_index(headers, ["备注", "note"])
    if status_idx is None:
        status_idx = len(headers)
        sheet.cell(1, status_idx + 1).value = status_header
    if note_idx is None:
        note_idx = max(len(headers), status_idx + 1)
        sheet.cell(1, note_idx + 1).value = "备注"

    row_number = int(payload.get("row") or 0)
    if not row_number:
        code = text(payload.get("code") or payload.get("source_code") or payload.get("target_code"))
        code_idx = header_index(headers, ["课件编码", "courseware_code", "code"])
        if code_idx is not None:
            for row in range(2, sheet.max_row + 1):
                if text(sheet.cell(row, code_idx + 1).value) == code:
                    row_number = row
                    break
    if not row_number:
        return

    ok = bool(payload.get("ok"))
    dry_run = bool(payload.get("dryRun"))
    if kind == "cover":
        status = "检查通过" if ok and dry_run else "已上传" if ok else "失败"
    elif kind == "relation":
        status = "检查通过" if ok and dry_run else "已关联" if ok else "失败"
    else:
        status = "检查通过" if ok and dry_run else "已复制" if ok else "失败"

    note = payload.get("error") or payload.get("note") or ""
    if ok and kind == "cover" and not dry_run:
        note = f"attachment_id={payload.get('attachmentId') or ''}; url={payload.get('url') or ''}"
    elif ok and kind == "relation":
        note = f"source_id={payload.get('source_id') or ''}; target_id={payload.get('target_id') or ''}; work_id={payload.get('work_id') or ''}"
    elif ok and kind == "resource_copy":
        note = f"source_id={payload.get('source_id') or ''}; target_id={payload.get('target_id') or ''}"

    sheet.cell(row_number, status_idx + 1).value = status
    sheet.cell(row_number, note_idx + 1).value = note
    workbook.save(CONFIG_FILE)


class ReusableThreadingHTTPServer(ThreadingHTTPServer):
    allow_reuse_address = True
    daemon_threads = True


class Handler(BaseHTTPRequestHandler):
    server_version = "VipthinkPrelaunchAssistant/1.0 By Haihan"

    def end_headers(self) -> None:
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        super().end_headers()

    def do_OPTIONS(self) -> None:
        self.send_response(204)
        self.end_headers()

    def send_json(self, payload: object, status: int = 200) -> None:
        data = json_bytes(payload)
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/health":
                self.send_json({
                    "ok": True,
                    "author": "海瀚",
                    "root": str(ROOT),
                    "port": PORT,
                    "config_file": str(CONFIG_FILE),
                    "courseware_data_dir": str(COURSEWARE_DATA_DIR),
                })
                return
            if parsed.path == "/shutdown":
                self.send_json({"ok": True, "message": "assistant is shutting down"})
                threading.Thread(target=self.server.shutdown, daemon=True).start()
                return
            if parsed.path == "/tasks":
                selected = (parse_qs(parsed.query).get("config") or [""])[0]
                self.send_json({"ok": True, "tasks": build_cover_tasks(resolve_workbook_path("prelaunch", selected))})
                return
            if parsed.path == "/relation-tasks":
                selected = (parse_qs(parsed.query).get("config") or [""])[0]
                tasks = build_relation_tasks(resolve_workbook_path("prelaunch", selected))
                self.send_json({"ok": True, "tasks": tasks, "count": len(tasks)})
                return
            if parsed.path == "/resource-copy-tasks":
                selected = (parse_qs(parsed.query).get("config") or [""])[0]
                tasks = build_resource_copy_tasks(resolve_workbook_path("prelaunch", selected))
                self.send_json({"ok": True, "tasks": tasks, "count": len(tasks)})
                return
            if parsed.path == "/workbook-candidates":
                query = parse_qs(parsed.query)
                kind = (query.get("kind") or ["prelaunch"])[0]
                self.send_json({"ok": True, "candidates": workbook_candidates(kind)})
                return
            if parsed.path == "/knowledge-copy-tasks":
                selected = (parse_qs(parsed.query).get("config") or [""])[0]
                tasks = build_knowledge_copy_tasks(resolve_workbook_path("postlaunch", selected))
                self.send_json({"ok": True, "tasks": tasks, "count": len(tasks)})
                return
            if parsed.path == "/small-teacher-upload-tasks":
                selected = (parse_qs(parsed.query).get("config") or [""])[0]
                tasks = build_small_teacher_upload_tasks(resolve_workbook_path("small_teacher", selected))
                self.send_json({"ok": True, "tasks": tasks, "count": len(tasks)})
                return
            if parsed.path == "/resource-copy-template":
                self.send_json({"ok": True, "template": load_resource_copy_template()})
                return
            if parsed.path == "/template-options":
                self.send_json({"ok": True, "options": template_options()})
                return
            if parsed.path == "/image":
                query = parse_qs(parsed.query)
                code = (query.get("code") or [""])[0]
                task = next((item for item in build_cover_tasks() if item["courseware_code"] == code), None)
                if not task or not task.get("image_exists"):
                    self.send_json({"ok": False, "error": f"找不到图片：{code}"}, 404)
                    return
                image_path = Path(task["image_path"])
                content_type = mimetypes.guess_type(image_path.name)[0] or "application/octet-stream"
                data = image_path.read_bytes()
                self.send_response(200)
                self.send_header("Content-Type", content_type)
                self.send_header("Content-Length", str(len(data)))
                self.send_header("X-File-Name", image_path.name)
                self.end_headers()
                self.wfile.write(data)
                return
            if parsed.path == "/small-teacher-image":
                query = parse_qs(parsed.query)
                row = (query.get("row") or [""])[0]
                code = (query.get("code") or [""])[0]
                task = small_teacher_task_by_row_or_code(row, code)
                image_path = find_small_teacher_image(task.get("image", "") if task else "", code)
                if not image_path:
                    self.send_json({
                        "ok": False,
                        "error": f"找不到小老师图片：{code or row}",
                        "image_dir": str(SMALL_TEACHER_IMAGE_DIR),
                    }, 404)
                    return
                content_type = mimetypes.guess_type(image_path.name)[0] or "application/octet-stream"
                data = image_path.read_bytes()
                self.send_response(200)
                self.send_header("Content-Type", content_type)
                self.send_header("Content-Length", str(len(data)))
                self.send_header("X-File-Name", image_path.name)
                self.end_headers()
                self.wfile.write(data)
                return
        except Exception as error:
            self.send_json({"ok": False, "error": str(error)}, 500)
            return
        self.send_json({"ok": False, "error": "Not found"}, 404)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        length = int(self.headers.get("Content-Length") or 0)
        body = self.rfile.read(length).decode("utf-8") if length else "{}"
        try:
            payload = json.loads(body)
        except json.JSONDecodeError:
            payload = {}

        if parsed.path == "/shutdown":
            self.send_json({"ok": True, "message": "assistant is shutting down"})
            threading.Thread(target=self.server.shutdown, daemon=True).start()
            return

        kind_by_path = {
            "/result": "cover",
            "/relation-result": "relation",
            "/resource-copy-result": "resource_copy",
        }
        if parsed.path == "/courseware-data-result":
            try:
                output_path = write_courseware_data(payload)
            except Exception as error:
                self.send_json({"ok": False, "error": str(error)}, 500)
                return
            self.send_json({"ok": True, "path": str(output_path), "count": len(payload.get("rows") or [])})
            return
        if parsed.path == "/small-teacher-result":
            try:
                result = write_small_teacher_data(payload)
            except Exception as error:
                self.send_json({"ok": False, "error": str(error)}, 500)
                return
            self.send_json({"ok": True, **result})
            return
        if parsed.path == "/fill-template":
            try:
                result = fill_latest_data_to_template(payload)
            except Exception as error:
                self.send_json({"ok": False, "error": str(error)}, 500)
                return
            self.send_json({"ok": True, **result})
            return

        kind = kind_by_path.get(parsed.path)
        if kind:
            append_result(kind, payload)
            update_sheet_result(kind, payload)
            self.send_json({"ok": True})
            return
        self.send_json({"ok": False, "error": "Not found"}, 404)

    def log_message(self, fmt: str, *args) -> None:
        print(fmt % args)


def main() -> None:
    ensure_config_file()
    ensure_postlaunch_config_file()
    ensure_small_teacher_config_file()
    SMALL_TEACHER_IMAGE_DIR.mkdir(exist_ok=True)
    print(f"课件上架前配置本地助手已启动：http://{HOST}:{PORT}")
    print(f"作者：海瀚")
    print(f"统一配置表：{CONFIG_FILE}")
    print("请保持窗口打开，然后在 Chrome/Edge 插件中执行检查或配置。")
    server = ReusableThreadingHTTPServer((HOST, PORT), Handler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()
        print("课件上架前配置本地助手已关闭，端口已释放。")


if __name__ == "__main__":
    main()
