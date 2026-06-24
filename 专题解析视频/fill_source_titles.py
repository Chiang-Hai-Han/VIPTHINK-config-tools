from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
SHEET_FILE = ROOT / "专题解析视频配置表.xlsx"
MAPPING_FILE = ROOT / "字符映射.json"
SHEET_NAME = "专题解析视频"


def text(value) -> str:
    return "" if value is None else str(value).strip()


def load_mappings() -> list[dict[str, str]]:
    if not MAPPING_FILE.exists():
        raise FileNotFoundError(f"找不到映射文件：{MAPPING_FILE}")
    data = json.loads(MAPPING_FILE.read_text(encoding="utf-8"))
    rules = data.get("replace_rules") or []
    normalized: list[dict[str, str]] = []
    for item in rules:
        if not isinstance(item, dict):
            continue
        source = text(item.get("target_contains"))
        dest = text(item.get("source_contains"))
        if source and dest:
            normalized.append({"target_contains": source, "source_contains": dest})
    if not normalized:
        raise RuntimeError("字符映射.json 里没有可用的 replace_rules。")
    return normalized


def extract_question_no(title: str) -> str:
    match = re.search(r"第\s*(\d+)\s*题", title)
    return match.group(1) if match else ""


def apply_mapping(target_title: str, rules: list[dict[str, str]]) -> str:
    for rule in rules:
        if rule["target_contains"] in target_title:
            return target_title.replace(rule["target_contains"], rule["source_contains"])
    return target_title


def main() -> None:
    if not SHEET_FILE.exists():
        raise FileNotFoundError(f"找不到配置表：{SHEET_FILE}")

    rules = load_mappings()
    workbook = openpyxl.load_workbook(SHEET_FILE)
    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active

    headers = {text(sheet.cell(1, col).value): col for col in range(1, sheet.max_column + 1)}
    required = ["目标题目名称", "来源题目名称", "目标课件编码", "来源课件编码", "题号"]
    for key in required:
        if key not in headers:
            raise RuntimeError(f"配置表缺少表头：{key}")

    updated = 0
    for row in range(2, sheet.max_row + 1):
        target_title = text(sheet.cell(row, headers["目标题目名称"]).value)
        if not target_title:
            continue
        source_title = apply_mapping(target_title, rules)
        sheet.cell(row, headers["来源题目名称"]).value = source_title

        question_no = extract_question_no(target_title)
        if question_no:
            sheet.cell(row, headers["题号"]).value = question_no

        target_code = re.sub(r"第\s*\d+\s*题.*$", "", target_title).strip()
        source_code = re.sub(r"第\s*\d+\s*题.*$", "", source_title).strip()
        sheet.cell(row, headers["目标课件编码"]).value = target_code
        sheet.cell(row, headers["来源课件编码"]).value = source_code
        updated += 1

    workbook.save(SHEET_FILE)
    print(f"已更新 {updated} 行：{SHEET_FILE}")


if __name__ == "__main__":
    main()
