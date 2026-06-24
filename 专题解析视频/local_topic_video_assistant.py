from __future__ import annotations

import json
import os
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import openpyxl


ROOT = Path(__file__).resolve().parent
CONFIG_FILE = ROOT / "专题解析视频配置表.xlsx"
SHEET_NAME = "专题解析视频"
HOST = "127.0.0.1"
PORT = int(os.environ.get("VIPTHINK_TOPIC_VIDEO_ASSISTANT_PORT", "8771"))
HEADERS = [
    "目标题目名称",
    "来源题目名称",
    "目标课件编码",
    "来源课件编码",
    "题号",
    "指导视频复制状态",
    "备注",
]
SAMPLE_ROW = [
    "s1_v7_01-04_TW第1题",
    "s1_v8_01-04_YY第1题",
    "s1_v7_01-04_TW",
    "s1_v8_01-04_YY",
    "1",
    "待处理",
    "示例：下载 v8 YY 的第1题指导视频，上传到 v7 TW 的第1题",
]


def json_bytes(payload: object) -> bytes:
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def text(value) -> str:
    return "" if value is None else str(value).strip()


def normalize_question_title(title: str, code: str, question_no: str) -> str:
    raw_title = text(title)
    raw_code = text(code)
    raw_no = text(question_no)
    if raw_title and "?" not in raw_title:
        return raw_title
    if raw_code and raw_no:
        return f"{raw_code}第{raw_no}题"
    return raw_title


def ensure_sheet() -> None:
    if CONFIG_FILE.exists():
        workbook = openpyxl.load_workbook(CONFIG_FILE)
        sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
        sheet.title = SHEET_NAME
        existing = [text(cell.value) for cell in sheet[1]]
        changed = False
        for col, header in enumerate(HEADERS, start=1):
            if col > len(existing) or existing[col - 1] != header:
                sheet.cell(1, col).value = header
                changed = True
        if sheet.max_row < 2:
            sheet.append(SAMPLE_ROW)
            changed = True
        if changed:
            workbook.save(CONFIG_FILE)
        return

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(HEADERS)
    sheet.append(SAMPLE_ROW)
    workbook.save(CONFIG_FILE)


def workbook_headers(sheet) -> dict[str, int]:
    return {
        text(sheet.cell(1, column).value): column
        for column in range(1, sheet.max_column + 1)
        if text(sheet.cell(1, column).value)
    }


def load_tasks() -> list[dict[str, object]]:
    ensure_sheet()
    workbook = openpyxl.load_workbook(CONFIG_FILE, data_only=True)
    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
    headers = workbook_headers(sheet)
    tasks: list[dict[str, object]] = []
    for row_number in range(2, sheet.max_row + 1):
        target_code = text(sheet.cell(row_number, headers["目标课件编码"]).value)
        source_code = text(sheet.cell(row_number, headers["来源课件编码"]).value)
        question_no = text(sheet.cell(row_number, headers["题号"]).value)
        target_title = normalize_question_title(
            text(sheet.cell(row_number, headers["目标题目名称"]).value),
            target_code,
            question_no,
        )
        source_title = normalize_question_title(
            text(sheet.cell(row_number, headers["来源题目名称"]).value),
            source_code,
            question_no,
        )
        if not target_title or not source_title:
            continue
        tasks.append({
            "row": row_number,
            "target_title": target_title,
            "source_title": source_title,
            "target_code": target_code,
            "source_code": source_code,
            "question_no": question_no,
            "status": text(sheet.cell(row_number, headers["指导视频复制状态"]).value),
            "note": text(sheet.cell(row_number, headers["备注"]).value),
        })
    return tasks


def update_result(payload: dict) -> dict[str, object]:
    ensure_sheet()
    workbook = openpyxl.load_workbook(CONFIG_FILE)
    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
    headers = workbook_headers(sheet)

    row_number = int(payload.get("row") or 0)
    if row_number < 2 or row_number > sheet.max_row:
        raise RuntimeError(f"无效行号：{row_number}")

    status_col = headers.get("指导视频复制状态")
    note_col = headers.get("备注")
    if not status_col or not note_col:
        raise RuntimeError("配置表缺少状态或备注列。")

    sheet.cell(row_number, status_col).value = text(payload.get("status"))
    sheet.cell(row_number, note_col).value = text(payload.get("note"))
    workbook.save(CONFIG_FILE)
    return {"ok": True, "row": row_number}


class Handler(BaseHTTPRequestHandler):
    server_version = "TopicVideoAssistant/1.0"

    def log_message(self, format: str, *args) -> None:
        return

    def end_headers(self) -> None:
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        super().end_headers()

    def do_OPTIONS(self) -> None:
        self.send_response(204)
        self.end_headers()

    def send_json(self, payload: object, status: int = 200) -> None:
        data = json_bytes(payload)
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/health":
                self.send_json({"ok": True, "service": "topic-video-assistant", "config": str(CONFIG_FILE)})
                return
            if parsed.path == "/topic-video-tasks":
                query = parse_qs(parsed.query)
                include_done = text((query.get("include_done") or [""])[0]).lower() in {"1", "true", "yes"}
                tasks = load_tasks()
                if not include_done:
                    tasks = [task for task in tasks if text(task.get("status")) not in {"已完成", "跳过"}]
                self.send_json({"ok": True, "tasks": tasks})
                return
            self.send_json({"ok": False, "error": f"未知路径：{parsed.path}"}, 404)
        except Exception as error:
            self.send_json({"ok": False, "error": str(error)}, 500)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        length = int(self.headers.get("Content-Length") or 0)
        body = self.rfile.read(length).decode("utf-8") if length else "{}"
        try:
            payload = json.loads(body)
        except json.JSONDecodeError:
            self.send_json({"ok": False, "error": "请求体不是合法 JSON。"}, 400)
            return

        try:
            if parsed.path == "/topic-video-result":
                self.send_json(update_result(payload))
                return
            self.send_json({"ok": False, "error": f"未知路径：{parsed.path}"}, 404)
        except Exception as error:
            self.send_json({"ok": False, "error": str(error)}, 500)


def main() -> None:
    ensure_sheet()
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    print(f"专题解析视频本地助手已启动：http://{HOST}:{PORT}")
    print(f"配置表：{CONFIG_FILE}")
    server.serve_forever()


if __name__ == "__main__":
    main()
