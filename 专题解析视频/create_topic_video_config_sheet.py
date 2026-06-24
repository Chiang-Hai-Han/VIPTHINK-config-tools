from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
SHEET_FILE = ROOT / "专题解析视频配置表.xlsx"
SHEET_NAME = "专题解析视频"
HEADERS = [
    "目标题目名称",
    "来源题目名称",
    "目标课件编码",
    "来源课件编码",
    "题号",
    "指导视频复制状态",
    "备注",
]
EXAMPLE = [
    "s1_v7_01-04_TW第1题",
    "s1_v8_01-04_YY第1题",
    "s1_v7_01-04_TW",
    "s1_v8_01-04_YY",
    "1",
    "待处理",
    "示例：下载 v8 YY 的第1题指导视频，上传到 v7 TW 的第1题",
]


def ensure_sheet() -> Path:
    if SHEET_FILE.exists():
        workbook = openpyxl.load_workbook(SHEET_FILE)
        sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
        sheet.title = SHEET_NAME
        existing_headers = [str(cell.value or "").strip() for cell in sheet[1]]
        for col, header in enumerate(HEADERS, start=1):
            if col > len(existing_headers) or existing_headers[col - 1] != header:
                sheet.cell(1, col).value = header
        if sheet.max_row < 2:
            sheet.append(EXAMPLE)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAME
        sheet.append(HEADERS)
        sheet.append(EXAMPLE)

    widths = [28, 28, 20, 20, 10, 18, 48]
    for col, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    workbook.save(SHEET_FILE)
    return SHEET_FILE


def main() -> None:
    print(ensure_sheet())


if __name__ == "__main__":
    main()
