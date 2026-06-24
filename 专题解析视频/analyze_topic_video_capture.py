from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = ROOT / "监听结果"
MAPPING_FILE = ROOT / "字符映射.json"
CONFIG_FILE = ROOT / "专题解析视频配置表.xlsx"


def text(value) -> str:
    return "" if value is None else str(value)


def latest_capture_file() -> Path:
    files = sorted(OUTPUT_DIR.glob("专题解析视频接口监听_*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"监听结果目录中没有找到文件：{OUTPUT_DIR}")
    return files[0]


def flatten_json(value, prefix: str = ""):
    if isinstance(value, dict):
        for key, child in value.items():
            child_prefix = f"{prefix}.{key}" if prefix else str(key)
            yield from flatten_json(child, child_prefix)
    elif isinstance(value, list):
        for index, child in enumerate(value):
            child_prefix = f"{prefix}[{index}]"
            yield from flatten_json(child, child_prefix)
    else:
        yield prefix, value


def load_sample_pairs() -> tuple[str, str]:
    if CONFIG_FILE.exists():
        try:
            workbook = openpyxl.load_workbook(CONFIG_FILE, data_only=True)
            sheet = workbook.active
            target_title = text(sheet.cell(2, 1).value)
            source_title = text(sheet.cell(2, 2).value)
            if source_title and target_title:
                return source_title, target_title
        except Exception:
            pass
    if not MAPPING_FILE.exists():
        return "", ""
    data = json.loads(MAPPING_FILE.read_text(encoding="utf-8"))
    sample = data.get("sample_pair") or {}
    return text(sample.get("source_title")), text(sample.get("target_title"))


def looks_like_question_title(value: str) -> bool:
    return bool(re.search(r"s\d+_v\d+_.+第\s*\d+\s*题", value, flags=re.IGNORECASE))


def main() -> None:
    try:
        capture_file = latest_capture_file()
    except FileNotFoundError as error:
        print(str(error))
        print("请先运行 run_listen_topic_video_requests.bat，手动操作一次后再分析。")
        return
    sample_source, sample_target = load_sample_pairs()
    captures = json.loads(capture_file.read_text(encoding="utf-8"))

    print(f"分析文件：{capture_file}")
    print()

    total_candidates = 0
    for index, item in enumerate(captures, start=1):
        request = item.get("request") or {}
        payload_text = text(request.get("postData"))
        response_text = text(item.get("responsePreview"))
        joined = payload_text + "\n" + response_text

        score = 0
        if sample_source and sample_source in joined:
            score += 2
        if sample_target and sample_target in joined:
            score += 2
        if re.search(r"指导视频|video|upload|copy|attachment|resource", joined, flags=re.IGNORECASE):
            score += 1
        if score <= 0:
            continue

        total_candidates += 1
        print(f"[候选 {total_candidates}] {item.get('url', '')}")
        print(f"状态码：{item.get('status', '')}")

        try:
            request_json = json.loads(payload_text) if payload_text.strip().startswith(("{", "[")) else None
        except json.JSONDecodeError:
            request_json = None
        try:
            response_json = json.loads(response_text) if response_text.strip().startswith(("{", "[")) else None
        except json.JSONDecodeError:
            response_json = None

        printed = 0
        for label, data in [("request", request_json), ("response", response_json)]:
            if data is None:
                continue
            for path, value in flatten_json(data):
                value_text = text(value)
                if not value_text:
                    continue
                if sample_source and sample_source in value_text:
                    print(f"  {label} 命中来源题目：{path} = {value_text[:200]}")
                    printed += 1
                elif sample_target and sample_target in value_text:
                    print(f"  {label} 命中目标题目：{path} = {value_text[:200]}")
                    printed += 1
                elif looks_like_question_title(value_text):
                    print(f"  {label} 疑似题目字段：{path} = {value_text[:200]}")
                    printed += 1
                elif re.search(r"video|mp4|m3u8|oss|attachment|resource", value_text, flags=re.IGNORECASE):
                    print(f"  {label} 疑似视频字段：{path} = {value_text[:200]}")
                    printed += 1
                if printed >= 16:
                    break
            if printed >= 16:
                break

        if printed == 0:
            print("  没有自动识别出明显字段，可以直接打开监听文件查 postData / responsePreview。")
        print()

    if total_candidates == 0:
        print("没有筛到高相关接口。建议重新监听一次，并确保手动操作里包含：打开来源题目、复制指导视频、保存到目标题目。")


if __name__ == "__main__":
    main()
