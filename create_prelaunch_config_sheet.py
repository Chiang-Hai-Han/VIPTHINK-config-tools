from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
SHEET_FILE = ROOT / "课件上架前配置表.xlsx"
SHEET_NAME = "课件上架前配置"
HEADERS = [
    "课件编码",
    "课件名称",
    "封面图片",
    "作业课件编码",
    "资源来源课件",
    "封面上传状态",
    "课件关联状态",
    "资源复制状态",
    "备注",
]
EXAMPLE = [
    "s4_v8_04_TW",
    "",
    "",
    "s4_v8_04_TW_hw",
    "s4_v8_04_YY",
    "待上传",
    "待关联",
    "待复制",
    "示例：同一行维护封面、作业课件关联、资源复制的编码对应关系",
]


def main() -> None:
    if SHEET_FILE.exists():
        workbook = openpyxl.load_workbook(SHEET_FILE)
        sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
        sheet.title = SHEET_NAME
        existing_headers = [str(cell.value or "").strip() for cell in sheet[1]]
        for col, header in enumerate(HEADERS, start=1):
            if col > len(existing_headers) or existing_headers[col - 1] != header:
                sheet.cell(1, col).value = header
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAME
        sheet.append(HEADERS)
        sheet.append(EXAMPLE)

    widths = [20, 24, 24, 24, 24, 16, 16, 16, 56]
    for col, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    workbook.save(SHEET_FILE)
    print(SHEET_FILE)


if __name__ == "__main__":
    main()
